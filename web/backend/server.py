"""
JinTai AI Agent — Web Backend
FastAPI + tool calling (bom_parser, comparator) + streaming SSE
"""

import sys
import os
import json
import asyncio
import uuid
from pathlib import Path
from datetime import datetime

# Thêm project root vào path
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "agents"))
sys.path.insert(0, str(PROJECT_ROOT / "tools"))

# Load .env file
try:
    from dotenv import load_dotenv
    load_dotenv(str(PROJECT_ROOT / ".env"))
except ImportError:
    pass

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import BaseModel

# ── Import tools ──
from tools.bom_parser import parse_product_bom
from tools.comparator import compare_bom as _compare_bom

app = FastAPI(title="JinTai AI Agent")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

_sessions = {}
DATA_DIR = str(PROJECT_ROOT / "data")
SESSIONS_FILE = str(PROJECT_ROOT / "web" / "sessions.json")
MEMORY_FILE = str(PROJECT_ROOT / "web" / "memory.json")


# ── Memory persistence ──
def _load_memory():
    """Load memory from JSON file."""
    if os.path.exists(MEMORY_FILE):
        try:
            with open(MEMORY_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {"learned": []}


def _save_memory(memory):
    """Save memory to JSON file."""
    try:
        with open(MEMORY_FILE, 'w', encoding='utf-8') as f:
            json.dump(memory, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error saving memory: {e}")


# ── Session persistence ──
def _load_sessions():
    """Load sessions from JSON file."""
    if os.path.exists(SESSIONS_FILE):
        try:
            with open(SESSIONS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def _save_sessions():
    """Save sessions to JSON file."""
    try:
        with open(SESSIONS_FILE, 'w', encoding='utf-8') as f:
            json.dump(_sessions, f, ensure_ascii=False, indent=2, default=str)
    except Exception as e:
        print(f"Error saving sessions: {e}")


# Load on startup
_sessions = _load_sessions()


# ── Parse owl-alpha custom tool call format ──
def _parse_custom_tool_call(text: str) -> dict | None:
    """Parse <longcat_tool_call>...</longcat_tool_call> format."""
    import re
    match = re.search(
        r'<longcat_tool_call>\s*(\w+)\s*<longcat_arg_key>(\w+)</longcat_arg_key>\s*<longcat_arg_value>(.*?)</longcat_arg_value>\s*</longcat_tool_call>',
        text, re.DOTALL
    )
    if match:
        tool_name = match.group(1)
        arg_key = match.group(2)
        arg_val = match.group(3).strip()
        return {"name": tool_name, "args": {arg_key: arg_val}}

    # Try multi-arg format
    match = re.search(r'<longcat_tool_call>(.*?)</longcat_tool_call>', text, re.DOTALL)
    if match:
        inner = match.group(1)
        tool_name_match = re.match(r'\s*(\w+)', inner)
        if tool_name_match:
            tool_name = tool_name_match.group(1)
            args = {}
            for m in re.finditer(r'<longcat_arg_key>(\w+)</longcat_arg_key>\s*<longcat_arg_value>(.*?)</longcat_arg_value>', inner, re.DOTALL):
                args[m.group(1)] = m.group(2).strip()
            if args:
                return {"name": tool_name, "args": args}

    return None

# ── Tool definitions for LLM ──
TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "bom_parser",
            "description": "Parse BOM (Bill of Materials) CSV file for a product. Returns list of parts with material code, name, spec, color, category, component code, quantity.",
            "parameters": {
                "type": "object",
                "properties": {
                    "sku_id": {
                        "type": "string",
                        "description": "Product SKU code, e.g. 'LGS031'. Format: LGS + 3-4 digits."
                    },
                    "lang": {
                        "type": "string",
                        "enum": ["vi", "zh", "auto"],
                        "description": "Language: 'vi' for Vietnamese BOM, 'zh' for Chinese, 'auto' tries VI first."
                    }
                },
                "required": ["sku_id"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "comparator",
            "description": "Compare two product BOMs and find differences. Shows common parts, unique parts, quantity differences, and component code warnings.",
            "parameters": {
                "type": "object",
                "properties": {
                    "base_sku": {
                        "type": "string",
                        "description": "Base product SKU, e.g. 'LGS723'"
                    },
                    "compare_sku": {
                        "type": "string",
                        "description": "Product to compare against, e.g. 'LGS733'"
                    },
                    "lang": {
                        "type": "string",
                        "enum": ["vi", "zh", "auto"],
                        "description": "Language for BOM files."
                    },
                    "category": {
                        "type": "string",
                        "enum": ["all", "零件", "包材", "五金包", "铁件"],
                        "description": "Filter by category. Default 'all'."
                    }
                },
                "required": ["base_sku", "compare_sku"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "memory",
            "description": "Save or retrieve user preferences and learned facts. Use this to remember things about the user for future conversations.",
            "parameters": {
                "type": "object",
                "properties": {
                    "action": {
                        "type": "string",
                        "enum": ["add", "remove"],
                        "description": "Action: 'add' to save new memory, 'remove' to delete memory"
                    },
                    "content": {
                        "type": "string",
                        "description": "Memory content to save (in English). E.g., 'User name: An', 'User prefers full BOM display'"
                    }
                },
                "required": ["action", "content"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "export_excel",
            "description": "Export BOM data to Excel file (.xlsx). Returns download URL.",
            "parameters": {
                "type": "object",
                "properties": {
                    "sku_id": {
                        "type": "string",
                        "description": "Product SKU code, e.g. 'LGS031'"
                    },
                    "lang": {
                        "type": "string",
                        "enum": ["vi", "zh", "auto"],
                        "description": "Language for BOM files."
                    }
                },
                "required": ["sku_id"]
            }
        }
    }
]

SYSTEM_PROMPT = """You are JinTai AI Agent (金汰家具) for Jintai Furniture.
Built by Du Tuấn An (俞俊安).

═══ IDENTITY ═══
- Name: JinTai AI Agent
- Company: Jintai Furniture (金汰家具)
- Role: Internal BOM assistant
- Tone: Professional but friendly, concise, helpful. NO humor, NO poetry, NO flattery.

═══ ADDRESSING ═══
- Vietnamese: "Bạn" | Chinese: "你" | English: "you"

═══ LANGUAGE RULES ═══
1. ALWAYS reply in the SAME language the user writes in.
2. Vietnamese → reply Vietnamese. Chinese → reply Chinese. English → reply English.
3. BOM data contains Chinese terms. ALWAYS translate them to user's language.
4. NEVER output Chinese characters when user writes Vietnamese.

═══ SAFETY RULE ═══
- NEVER modify, delete, or overwrite any files. Only READ.

═══ SKU VALIDATION ═══
- Format: LGS + 3-4 digits (e.g., LGS031, LGS723)
- If user gives wrong format (like "402"), explain correct format.
- Available: LGS101,LGS111,LGS031,LGS032,LGS033,LGS043,LGS131,LGS132,LGS133,LGS231,LGS232,LGS233,LGS333,LGS334,LGS420,LGS421,LGS433,LGS434,LGS723,LGS733,LGS833,LGS834

═══ TOOL LANG RULE ═══
- User writes Vietnamese → pass lang="vi" to tools
- User writes Chinese → pass lang="zh" to tools
- Default lang="auto" reads VI first, fallback 中文

═══ THINK BEFORE ACTING ═══
- Think step by step before calling tools.
- Do NOT call tools blindly. Consider what the user actually needs.

═══ SECTION 1: PARSE — When to call bom_parser ═══
**Trigger:** User asks about parts/materials of a specific SKU.
  Examples: "BOM của LGS031", "LGS031 dùng vật liệu nào", "Check BOM LGS031"

**After calling bom_parser:**
- Check "status" FIRST. If "error" → show error_message, STOP. Do NOT guess.
- If fallback_warning exists → WARN user BEFORE showing table:
  "Không có đúng màu yêu cầu, đang dùng mặc định: {current_color}"
- Use the "summary" field DIRECTLY — it's already bilingual (VI/中文).
  Do NOT rewrite numbers or data from the summary to avoid errors.
- After showing BOM, you MUST tell user: (1) how many colors available (2) which color is shown.

═══ SECTION 2: COMPARE — When to call comparator ═══
**Trigger:** User compares 2 products.
  Examples: "LGS031 và LGS033 khác gì", "So sánh LGS031 vs LGS032"

**After calling comparator:**
- Show "summary" field EXACTLY as-is. Do NOT rewrite, summarize, or rephrase.
- Display as table or list. Do NOT change format or skip items.
- If 编号 differs → show the ⚠️ warning that's already in the summary.
- Do NOT use bom_parser to manually compare — use comparator directly.

═══ SECTION 3: DISPLAY RULES ═══
- items (list of dicts) → render as markdown table.
- Use column names from tool output as headers (do NOT rename columns).
- If > 15 rows → show first 15 + "... còn N chi tiết khác".
- If user asks by group (e.g., "chỉ xem phần điện"), filter items by 属性/thuộc tính field.

═══ RULES SUMMARY ═══
1. Validate SKU format before calling tools.
2. Check tool status before interpreting data.
3. Use summary field directly — do NOT rewrite numbers.
4. Translate Chinese terms to user's language.
5. Warn about fallback color before showing table.
6. Show comparator summary EXACTLY as-is.
7. Think step by step before acting.
8. NEVER say "I'm an AI" or "Great question!" — just answer directly.
9. Keep responses concise. No long explanations unless user asks.

═══ MEMORY RULES ═══
- When you learn something new about the user, call the memory tool to save it.
- ALWAYS write memory content in ENGLISH (for consistency across languages).
- Examples of what to remember:
  * User tells you their name → memory(add, "User name: X")
  * User says they prefer something → memory(add, "User prefers: Y")
  * User corrects you → memory(add, "Correction: X is actually Y")
  * User asks about same product multiple times → memory(add, "User frequently asks about: X")
- NEVER save sensitive data (passwords, tokens, API keys).
- Keep memories concise (1-2 sentences each).

═══ EDGE CASES ═══
- If user asks something OUTSIDE BOM/tools (greetings, thanks, general questions) → reply naturally in their language, no need to call tools.
- If user asks about a product NOT in the list → say "Sản phẩm [SKU] chưa có trong hệ thống" and show available products.
- If user wants to export data to Excel → call export_excel tool, then provide download link AND file path on D: drive.
- If tool returns "error" → show error_message, STOP. Do NOT guess or make up data."""


class ChatRequest(BaseModel):
    message: str
    session_id: str = None
    model: str = "openrouter/owl-alpha"
    provider: str = "openrouter"


class SessionData(BaseModel):
    id: str
    title: str
    created_at: str
    message_count: int
    preview: str = ""


# ── Execute tool ──
# ── Export Excel tool ──
EXPORT_DIR = str(PROJECT_ROOT / "web" / "exports")


def _copy_to_drive(filename: str) -> str:
    """Copy exported file to D: drive for easy access."""
    import shutil
    src = os.path.join(EXPORT_DIR, filename)
    if not os.path.exists(src):
        return None
    
    # Try WSL path first, then Windows path
    paths = [
        f"/mnt/d/{filename}",           # WSL
        f"D:\\{filename}",               # Windows
        os.path.expanduser(f"~/Desktop/{filename}"),  # Fallback
    ]
    
    for dst in paths:
        try:
            shutil.copy2(src, dst)
            return dst
        except:
            continue
    return None


def _export_excel(sku_id: str, lang: str = "auto") -> dict:
    """Export BOM data to Excel file with styling."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    # Get BOM data
    result = parse_product_bom(sku_id=sku_id, data_dir=DATA_DIR, lang=lang)
    if result.get("status") == "error":
        return {"status": "error", "error_message": result.get("error_message", "Unknown error")}

    # Create exports dir
    os.makedirs(EXPORT_DIR, exist_ok=True)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"BOM {sku_id}"

    # ── Styles ──
    header_fill = PatternFill(start_color="1F4E3D", end_color="1F4E3D", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # ── Headers (bilingual) ──
    headers = [
        ("物料编码", "Mã vật liệu"),
        ("物料名称", "Tên vật liệu"),
        ("规格型号", "Quy cách"),
        ("颜色", "Màu sắc"),
        ("属性", "Thuộc tính"),
        ("部件编号", "Mã linh kiện"),
        ("数量", "Số lượng"),
    ]

    for col, (zh, vi) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=f"{zh}\n{vi}")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 35  # Taller header for bilingual

    # ── Data rows ──
    items = result.get("items", [])
    for row_idx, item in enumerate(items, 2):
        values = [
            item.get("物料编码/mã vật liệu", ""),
            item.get("物料名称/tên vật liệu", ""),
            item.get("规格型号/quy cách", ""),
            item.get("颜色/màu sắc", ""),
            item.get("属性/thuộc tính", ""),
            item.get("部件编号/mã linh kiện", ""),
            item.get("数量/số lượng", 0),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if col == 7:  # Quantity
                cell.alignment = center_align

    # ── Summary row ──
    summary_row = len(items) + 3
    ws.cell(row=summary_row, column=1, value="Tổng số / 总数:").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=len(items)).font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value="Màu sắc / 颜色:").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=2, value=result.get("current_color", "")).font = Font(bold=True)

    # ── Column widths ──
    widths = [18, 35, 28, 15, 15, 18, 12]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Freeze header row ──
    ws.freeze_panes = 'A2'

    # ── Save ──
    filename = f"BOM_{sku_id}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)
    wb.close()

    # ── Copy to D: drive ──
    copied_path = _copy_to_drive(filename)

    return {
        "status": "success",
        "filename": filename,
        "download_url": f"/api/download/{filename}",
        "copied_to": copied_path,
        "total_items": len(items),
        "color": result.get("current_color", ""),
        "colors_available": result.get("available_colors", []),
    }


def execute_tool(name: str, args: dict) -> dict:
    if name == "bom_parser":
        return parse_product_bom(
            sku_id=args["sku_id"],
            data_dir=DATA_DIR,
            lang=args.get("lang", "auto"),
        )
    elif name == "comparator":
        result = _compare_bom(
            base_sku=args["base_sku"],
            compare_sku=args["compare_sku"],
            data_dir=DATA_DIR,
            lang=args.get("lang", "auto"),
            category=args.get("category", "all"),
        )
        # Simplify for LLM
        if result.get("error"):
            return {"status": "error", "error_message": result["message"]}
        return {
            "status": "success",
            "summary": result["summary"],
            "similarity": result.get("similarity_score", 0),
            "common_count": result.get("common_count", 0),
            "unique_base_count": result.get("unique_base_count", 0),
            "unique_compare_count": result.get("unique_compare_count", 0),
        }
    elif name == "memory":
        action = args.get("action", "add")
        content = args.get("content", "")
        memory = _load_memory()
        if action == "add" and content:
            if content not in memory["learned"]:
                memory["learned"].append(content)
                _save_memory(memory)
                return {"status": "saved", "content": content}
            return {"status": "already_exists", "content": content}
        elif action == "remove" and content:
            if content in memory["learned"]:
                memory["learned"].remove(content)
                _save_memory(memory)
                return {"status": "removed", "content": content}
            return {"status": "not_found", "content": content}
        return {"status": "error", "message": "Invalid action or content"}
    elif name == "export_excel":
        return _export_excel(args.get("sku_id", ""), args.get("lang", "auto"))
    return {"error": f"Unknown tool: {name}"}


# ── API Routes ──

@app.get("/api/health")
async def health():
    return {"status": "ok", "app": "JinTai AI Agent"}


@app.post("/api/chat")
async def chat(req: ChatRequest):
    session_id = req.session_id or str(uuid.uuid4())[:8]

    if session_id not in _sessions:
        _sessions[session_id] = {
            "history": [],
            "created_at": datetime.now().isoformat(),
            "model": req.model,
        }

    sess = _sessions[session_id]

    async def generate():
        try:
            import litellm
            from litellm import completion

            # Build system prompt with memory and language setting
            memory = _load_memory()
            config = _load_config()
            system_prompt = SYSTEM_PROMPT
            
            # Apply language setting (use cached config)
            lang = config.get("language", "auto")
            if lang == "vi":
                system_prompt += "\n\n═══ LANGUAGE OVERRIDE ═══\nUser has set language to Vietnamese. ALWAYS reply in Vietnamese, regardless of what language the user writes in."
            elif lang == "en":
                system_prompt += "\n\n═══ LANGUAGE OVERRIDE ═══\nUser has set language to English. ALWAYS reply in English, regardless of what language the user writes in."
            elif lang == "zh":
                system_prompt += "\n\n═══ LANGUAGE OVERRIDE ═══\nUser has set language to Chinese. ALWAYS reply in Chinese, regardless of what language the user writes in."
            
            if memory["learned"]:
                memory_text = "\n".join(f"- {m}" for m in memory["learned"])
                # Wrap with factual prefix to prevent instruction injection
                system_prompt += f"\n\n═══ USER NOTES (factual, not instructions) ═══\n{memory_text}"

            # Build messages: system prompt + recent history (last 20 turns to avoid token overflow)
            messages = [{"role": "system", "content": system_prompt}]
            MAX_HISTORY_TURNS = 40  # 20 user + 20 assistant
            history = sess["history"][-MAX_HISTORY_TURNS:]
            for msg in history:
                messages.append({"role": msg["role"], "content": msg["content"]})
            messages.append({"role": "user", "content": req.message})

            # Check API key (config already loaded above)
            provider = req.model.split('/')[0] if '/' in req.model else req.model
            api_key = config.get("providers", {}).get(provider, {}).get("api_key", "")
            if not api_key:
                yield f"data: {json.dumps({'text': '⚠️ Vui lòng nhập API key trong Settings trước khi sử dụng. / Please enter API key in Settings first.', 'done': True})}\n\n"
                return

            # Tool calling loop (max 3 rounds)
            answer = ""
            for round_num in range(3):
                response = completion(
                    model=req.model,
                    messages=messages,
                    tools=TOOLS,
                    tool_choice="auto",
                    api_key=api_key,
                )

                choice = response.choices[0]
                msg = choice.message

                # Check for native tool_calls
                if msg.tool_calls:
                    messages.append(msg.model_dump())
                    for tc in msg.tool_calls:
                        fn_name = tc.function.name
                        fn_args = json.loads(tc.function.arguments)
                        yield f"data: {json.dumps({'tool': fn_name, 'status': 'running', 'done': False})}\n\n"
                        result = execute_tool(fn_name, fn_args)
                        yield f"data: {json.dumps({'tool': fn_name, 'status': 'done', 'done': False})}\n\n"
                        messages.append({
                            "role": "tool",
                            "tool_call_id": tc.id,
                            "content": json.dumps(result, ensure_ascii=False, default=str)[:8000],
                        })
                    continue

                # Check for custom tool call format (owl-alpha style)
                raw = msg.content or ""
                tool_call = _parse_custom_tool_call(raw)
                if tool_call:
                    fn_name = tool_call["name"]
                    fn_args = tool_call["args"]
                    yield f"data: {json.dumps({'tool': fn_name, 'status': 'running', 'done': False})}\n\n"
                    result = execute_tool(fn_name, fn_args)
                    yield f"data: {json.dumps({'tool': fn_name, 'status': 'done', 'done': False})}\n\n"
                    # Add tool result as user message for next round
                    messages.append({"role": "assistant", "content": raw})
                    messages.append({"role": "user", "content": f"Tool {fn_name} result:\n{json.dumps(result, ensure_ascii=False, default=str)[:8000]}\n\nNow summarize this for the user in their language."})
                    continue

                # No tool call → final answer
                answer = raw
                messages.append({"role": "assistant", "content": answer})
                break

            # Stream final answer character by character
            for i in range(0, len(answer), 3):
                chunk = answer[i:i+3]
                yield f"data: {json.dumps({'text': chunk, 'done': False})}\n\n"
                await asyncio.sleep(0.01)

            # Save history
            sess["history"].append({"role": "user", "content": req.message})
            sess["history"].append({"role": "assistant", "content": answer})
            _save_sessions()

            yield f"data: {json.dumps({'done': True, 'session_id': session_id})}\n\n"

        except Exception as e:
            import traceback
            traceback.print_exc()
            yield f"data: {json.dumps({'error': str(e), 'done': True})}\n\n"

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
    )


@app.get("/api/sessions")
async def list_sessions():
    MAX_SESSIONS = 500  # Limit to avoid slow loading
    sessions = []
    for sid, data in _sessions.items():
        h = data["history"]
        preview = h[-1]["content"][:80] + "..." if h else ""
        title = h[0]["content"][:40] if h else "New Chat"
        sessions.append(SessionData(
            id=sid, title=title, created_at=data["created_at"],
            message_count=len(h) // 2, preview=preview,
        ))
    sessions.sort(key=lambda s: s.created_at, reverse=True)
    # Limit to most recent sessions
    sessions = sessions[:MAX_SESSIONS]
    return {"sessions": [s.model_dump() for s in sessions]}


@app.get("/api/sessions/{session_id}")
async def get_session(session_id: str):
    MAX_MESSAGES = 50  # Limit messages per session
    if session_id not in _sessions:
        raise HTTPException(404, "Session not found")
    history = _sessions[session_id]["history"]
    # Return only most recent messages
    return {"session_id": session_id, "messages": history[-MAX_MESSAGES:]}


@app.delete("/api/sessions/{session_id}")
async def delete_session(session_id: str):
    if session_id in _sessions:
        del _sessions[session_id]
        _save_sessions()
    return {"status": "deleted"}


# ── Config/Settings ──
CONFIG_FILE = str(PROJECT_ROOT / "web" / "config.json")


def _load_config():
    """Load config from JSON file."""
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {"providers": {}, "default_model": "openrouter/owl-alpha"}


def _save_config(config):
    """Save config to JSON file."""
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error saving config: {e}")


@app.get("/api/settings")
async def get_settings():
    """Get current settings (API keys masked)."""
    config = _load_config()
    # Mask API keys for security
    masked = json.loads(json.dumps(config))
    for provider in masked.get("providers", {}):
        key = masked["providers"][provider].get("api_key", "")
        if key and len(key) > 8:
            masked["providers"][provider]["api_key"] = key[:4] + "..." + key[-4:]
    return masked


@app.post("/api/settings")
async def update_settings(req: dict):
    """Update settings (API keys)."""
    config = _load_config()
    if "providers" in req:
        for provider, data in req["providers"].items():
            if provider not in config["providers"]:
                config["providers"][provider] = {"api_key": "", "models": []}
            if "api_key" in data and data["api_key"]:
                # Only update if not masked
                if "..." not in data["api_key"]:
                    config["providers"][provider]["api_key"] = data["api_key"]
            if "models" in data:
                config["providers"][provider]["models"] = data["models"]
    if "default_model" in req:
        config["default_model"] = req["default_model"]
    if "language" in req:
        config["language"] = req["language"]
    _save_config(config)
    return {"status": "saved"}


@app.get("/api/download/{filename}")
async def download_file(filename: str):
    """Download exported Excel file."""
    import re
    # Sanitize filename
    if not re.match(r'^[\w\u4e00-\u9fff\u00C0-\u024F\u0300-\u036F\u1E00-\u1EFF-]+\.xlsx$', filename):
        raise HTTPException(400, "Invalid filename")
    filepath = os.path.join(EXPORT_DIR, filename)
    if not os.path.exists(filepath):
        raise HTTPException(404, "File not found")
    return FileResponse(filepath, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/models")
async def list_models():
    return {
        "providers": [
            {
                "id": "openrouter",
                "name": "OpenRouter",
                "models": [
                    {"id": "openrouter/owl-alpha", "name": "OWL Alpha", "cost": "free"},
                    {"id": "openrouter/mimo-v2.5", "name": "MiMo V2.5", "cost": "$0.013/1M"},
                    {"id": "openrouter/deepseek-v4", "name": "DeepSeek V4", "cost": "$0.068/1M"},
                ],
            },
            {
                "id": "google",
                "name": "Google AI Studio (Gemini)",
                "models": [
                    {"id": "google/gemini-2.5-flash", "name": "Gemini 2.5 Flash", "cost": "free"},
                    {"id": "google/gemini-2.5-pro", "name": "Gemini 2.5 Pro", "cost": "tiered"},
                ],
            },
        ]
    }


# ── Serve frontend ──
FRONTEND_DIR = PROJECT_ROOT / "web" / "frontend" / "dist"


@app.get("/{full_path:path}")
async def serve_frontend(full_path: str):
    if full_path.startswith("api/"):
        raise HTTPException(404)
    if not FRONTEND_DIR.exists():
        return {"message": "Build frontend: cd web/frontend && npm run build"}
    file_path = FRONTEND_DIR / full_path
    if file_path.exists() and file_path.is_file():
        return FileResponse(str(file_path))
    index_path = FRONTEND_DIR / "index.html"
    if index_path.exists():
        return FileResponse(str(index_path))
    return {"error": "Frontend not built"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8001)
