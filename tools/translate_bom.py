#!/usr/bin/env python3
"""Dịch BOM file từ 中文 sang VI dùng dictionary terms_zh_vi.json"""
import json
import openpyxl
import copy
import sys
import os

# Load dictionary
DICT_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "terms_zh_vi.json")
with open(DICT_PATH, encoding="utf-8") as f:
    TERMS = json.load(f)

# Color name mapping (sheet names + cell values)
COLOR_MAP = {
    "复古色": "màu gỗ cổ",
    "黑色": "màu đen",
    "白色": "màu trắng",
    "黑砂纹": "đen nhám",
    "黑泊板": "đen bóng",
    "本色": "màu tự nhiên",
    "烟墨黑色": "màu đen khói",
    "黑色3列": "màu đen 3 cột",
}

# Header translations (row 5)
HEADER_MAP = {
    "序号": "STT",
    "物料编码": "mã vật liệu",
    "部件编号": "mã linh kiện",
    "物料名称": "tên vật liệu",
    "规格型号": "quy cách",
    "材质": "chất liệu",
    "颜色": "màu sắc",
    "数量": "số lượng",
    "单位": "đơn vị",
}

# Row 0-2 translations
ROW_HEADER_MAP = {
    "型号": "mã hình thể",
    "编号": "mã số",
    "材质": "chất liệu",
    "颜色": "màu sắc",
}

# Material translations
MATERIAL_MAP = {
    "无纺布": "vải không dệt",
    "ABS": "ABS",
    "PB": "PB",
    "Q195": "Q195",
    "PP": "PP",
    "PP&GF40": "PP&GF40",
    "MDF": "MDF",
    "MDF&无纺布": "MDF&vải không dệt",
    "FPCB": "FPCB",
}


def translate_text(text, terms):
    """Translate Chinese text to Vietnamese using dictionary."""
    if not text or not isinstance(text, str):
        return text
    
    result = text
    
    # Try full match first
    if result in terms:
        return terms[result]
    
    # Try replacing known terms in the text
    # Sort by length (longest first) to avoid partial matches
    sorted_terms = sorted(terms.items(), key=lambda x: len(x[0]), reverse=True)
    
    for zh, vi in sorted_terms:
        if zh in result:
            result = result.replace(zh, vi)
    
    return result


def translate_color(text):
    """Translate color names."""
    if not text or not isinstance(text, str):
        return text
    
    # Check exact match in color map
    for zh, vi in COLOR_MAP.items():
        if zh in text:
            text = text.replace(zh, vi)
    
    # Also try dictionary
    if text in TERMS:
        return TERMS[text]
    
    return text


def translate_material(text):
    """Translate material names."""
    if not text or not isinstance(text, str):
        return text
    
    # Check material map
    for zh, vi in MATERIAL_MAP.items():
        if zh in text:
            text = text.replace(zh, vi)
    
    # Also try dictionary
    if text in TERMS:
        return TERMS[text]
    
    return text


def translate_sheet_name(name):
    """Translate sheet (color) name."""
    for zh, vi in COLOR_MAP.items():
        if zh in name:
            return name.replace(zh, vi)
    return name


def translate_bom(input_path, output_path):
    """Translate a BOM file from 中文 to VI."""
    wb = openpyxl.load_workbook(input_path)
    
    # Translate sheet names
    for sheet_name in wb.sheetnames:
        new_name = translate_sheet_name(sheet_name)
        if new_name != sheet_name:
            wb[sheet_name].title = new_name
    
    # Translate each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value is None or not isinstance(cell.value, str):
                    continue
                
                val = cell.value
                row_num = cell.row
                
                # Row 0: Title - translate color name
                if row_num == 1:
                    cell.value = translate_color(val)
                
                # Row 1: Dimensions - translate label
                elif row_num == 2:
                    for zh, vi in ROW_HEADER_MAP.items():
                        val = val.replace(f"{zh}:", f"{vi}:")
                        val = val.replace(f"{zh}：", f"{vi}：")
                    cell.value = val
                
                # Row 2: Product info - translate labels
                elif row_num == 3:
                    for zh, vi in ROW_HEADER_MAP.items():
                        val = val.replace(f"{zh}:", f"{vi}:")
                        val = val.replace(f"{zh}：", f"{vi}：")
                    # Also translate 布抽 in value
                    val = val.replace("布抽", "túi vải")
                    cell.value = val
                
                # Row 5 (header row): Translate column headers
                elif row_num == 6:
                    if val in HEADER_MAP:
                        cell.value = HEADER_MAP[val]
                
                # Data rows (6+): Translate specific columns
                elif row_num > 6:
                    col_letter = cell.column_letter
                    
                    # Column D: Part name (物料名称)
                    if col_letter == "D":
                        cell.value = translate_text(val, TERMS)
                    
                    # Column F: Material (材质)
                    elif col_letter == "F":
                        cell.value = translate_material(val)
                    
                    # Column G: Color (颜色)
                    elif col_letter == "G":
                        cell.value = translate_color(val)
    
    wb.save(output_path)
    print(f"✅ Translated: {input_path} → {output_path}")
    print(f"   Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python translate_bom.py <input.xlsx> [output.xlsx]")
        print("Example: python translate_bom.py data/sample_bom/BOM_LGS101_中文.xlsx")
        sys.exit(1)
    
    input_path = sys.argv[1]
    
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        # Auto-generate output path
        base = os.path.splitext(input_path)[0]
        output_path = base.replace("_中文", "_VI") + ".xlsx"
    
    translate_bom(input_path, output_path)
